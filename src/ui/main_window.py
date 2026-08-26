from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

from difflib import SequenceMatcher

from PySide6.QtCore import (
    Qt,
    QSortFilterProxyModel,
)

from PySide6.QtGui import (
    QIcon,
)

import qtawesome as qta

from PySide6.QtWidgets import (
    QMainWindow,
    QFileDialog,
    QMessageBox,
    QInputDialog,
    QLabel,
    QComboBox,
    QApplication,
    QHeaderView,
    QProgressDialog,
)

from PySide6.QtCore import QThread, Signal

from src.ui.themes import TEMAS
from src.ui.modelo_cambios import invalidar_cache_tema

from src.database.repositorio import RepositorioSQLite

from qextrawidgets.widgets.views import QFilterableTableView

from src.ui.ui_main_window import Ui_MainWindow

from src.ui.modelo_historico import ModeloHistorico

from src.database.repositorio_configuracion import (
    RepositorioConfiguracion,
    obtener_ruta_configuracion,
)

from src.excel.firma import generar_firma_excel

from src.excel.lector import (
    obtener_hojas,
    leer_excel,
)

from src.excel.nombres import (
    obtener_nombre_bd,
    obtener_identificador_propuesto,
)

from src.excel.analizador import (
    obtener_columnas_comunes,
    obtener_columnas_nuevas,
    obtener_columnas_eliminadas,
)

from src.ui.modelo_cambios import ModeloCambios

class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # =========================================================
        # TABLAS FILTRABLES
        # =========================================================

        self._crear_tabla_filtrable(
            "tableViewCambios"
        )

        self._crear_tabla_filtrable(
            "tableViewHistorico"
        )

        # =========================================================
        # MODELOS
        # =========================================================

        self.modelo_cambios = ModeloCambios()

        self.ui.tableViewCambios.setModel(
            self.modelo_cambios
        )

        self.modelo_historico = ModeloHistorico()

        self.ui.tableViewHistorico.setModel(
            self.modelo_historico
        )

        self.ui.lineEditBuscarClave.textChanged.connect(
            lambda texto: self._filtrar_lista(
                self.ui.listaColumnasClave,
                texto,
            )
        )

        self.ui.lineEditBuscarComparar.textChanged.connect(
            lambda texto: self._filtrar_lista(
                self.ui.listaColumnasComparar,
                texto,
            )
        )

        # DataFrames actualmente cargados
        self.df_1 = None
        self.df_2 = None

        # =========================================================
        # CONFIGURACIÓN EN CACHÉ
        # =========================================================

        self.repositorio_configuracion = (
            RepositorioConfiguracion(
                obtener_ruta_configuracion()
            )
        )

        self.firma_configuracion_actual = None
    
        self._conectar_eventos()

        self._inicializar_interfaz()

        self._configurar_contenido_ayuda()

        self._configurar_statusbar_y_temas()

        self._configurar_iconos()

        self._configurar_ventana_y_tablas()

        self.ui.tabWidget.currentChanged.connect(self._actualizar_iconos_pestanas)
        self._actualizar_iconos_pestanas(self.ui.tabWidget.currentIndex())

    def _configurar_contenido_ayuda(self):
        """
        Configura el texto explicativo y de contacto en el QTextBrowser
        de la pestaña 'Ayuda'.
        """

        html_contenido = """
        <div style="font-family: 'Segoe UI', sans-serif; padding: 10px;">
            <h2 style="color: #106EBE; margin-bottom: 5px;">📘 Guía de Uso del Comparador de Excel</h2>
            <p>Bienvenido al <b>Comparador de Excel</b>. Esta herramienta te permite comparar versiones de archivos Excel (como listas de cables F110) y llevar un registro organizado de los cambios.</p>

            <hr style="border: 0; height: 1px; background: #DCE3EC; margin: 15px 0;">

            <h3 style="color: #0F4C81;">🚀 Comparación Individual (2 archivos)</h3>
            <ol style="line-height: 1.6;">
                <li><b>Pestaña "Comparación individual":</b> Selecciona el <b>Excel 1</b> (versión anterior) y el <b>Excel 2</b> (versión nueva).</li>
                <li><b>Hojas de cálculo:</b> Selecciona la hoja correspondiente para cada archivo.</li>
                <li><b>Columnas clave:</b> Marca las columnas que identifican únicamente cada registro (por ejemplo: <i>Código, ID de Cable, Tag</i>).</li>
                <li><b>Columnas a comparar:</b> Selecciona las columnas cuyos valores deseas inspeccionar en búsqueda de diferencias.</li>
                <li><b>Comparar:</b> Haz clic en <b>"Comparar archivos"</b> para ver los resultados inmediatamente.</li>
            </ol>

            <h3 style="color: #0F4C81;">📦 Comparación Múltiple / Lote (Secuencial)</h3>
            <ol style="line-height: 1.6;">
                <li><b>Pestaña "Comparación múltiple":</b> Haz clic en <b>"Agregar archivos"</b> para cargar un grupo de archivos Excel organizados cronológicamente o por versiones.</li>
                <li><b>Reordenar secuencia:</b> Usa los botones ⬆/⬇ para asegurarte de que estén en el orden correcto (ej. <i>Versión 1 → Versión 2 → Versión 3</i>).</li>
                <li><b>Configuración común:</b> Selecciona la hoja común a comparar, las columnas clave y las columnas a evaluar.</li>
                <li><b>Procesar:</b> Haz clic en <b>"Procesar secuencialmente y crear histórico"</b> para generar automáticamente la base de datos de histórico con todas las comparaciones por parejas.</li>
            </ol>

            <h3 style="color: #0F4C81;">💾 Guardar e Histórico</h3>
            <ul style="line-height: 1.6;">
                <li>En la pestaña <b>"Resultados"</b>, haz clic en <b>"Guardar comparación"</b> para almacenar los cambios detectados en una base de datos SQLite.</li>
                <li>En la pestaña <b>"Histórico"</b>, puedes abrir una base de datos SQLite guardada para consultar comparaciones anteriores o exportar el histórico completo a Excel.</li>
            </ul>

            <hr style="border: 0; height: 1px; background: #DCE3EC; margin: 15px 0;">

            <h3 style="color: #106EBE;">✉️ Soporte y Contacto</h3>
            <p>Si tienes alguna sugerencia, duda o necesitas ayuda con la aplicación, ponte en contacto con:</p>
            <div style="background-color: #EBF0F5; border-left: 4px solid #106EBE; padding: 12px; border-radius: 4px; margin-top: 10px;">
                <b>Cristina Caravaca</b><br>
                Correo electrónico: <a href="mailto:cristina.caravaca@pine.zimacorp.es" style="color: #106EBE; font-weight: bold;">cristina.caravaca@pine.zimacorp.es</a>
            </div>
        </div>
        """

        self.ui.textBrowserAyuda.setHtml(html_contenido)

    # =========================================================
    # INICIALIZACIÓN
    # =========================================================

    def _inicializar_interfaz(self):
        """
        Configuración inicial de la interfaz.
        """

        self.ui.listaColumnasClave.clear()
        self.ui.listaColumnasComparar.clear()

        self.ui.labelResumen.setText(
            "0 cambios encontrados"
        )

    def _configurar_statusbar_y_temas(self):
        """
        Configura el indicador de base de datos activa y el selector
        de temas en la StatusBar.
        """

        # Indicador DB activa (izquierda)
        self.label_estado_db = QLabel("⚪ Sin base de datos cargada")
        self.ui.statusBar.addWidget(self.label_estado_db, 1)

        # Selector de Tema (derecha)
        label_tema = QLabel("Tema: ")
        self.combo_tema = QComboBox()
        self.combo_tema.addItems(list(TEMAS.keys()))

        # Cargar tema guardado
        tema_guardado = self.repositorio_configuracion.obtener_preferencia(
            "tema",
            "Pine Azul (Claro)",
        )

        if tema_guardado in TEMAS:
            self.combo_tema.setCurrentText(tema_guardado)
            self.cambiar_tema(tema_guardado)

        self.combo_tema.currentTextChanged.connect(self.cambiar_tema)

        self.ui.statusBar.addPermanentWidget(label_tema)
        self.ui.statusBar.addPermanentWidget(self.combo_tema)

    def cambiar_tema(self, nombre_tema: str):
        """
        Aplica el estilo QSS correspondiente al tema seleccionado
        y guarda la preferencia.
        """

        if nombre_tema in TEMAS:
            app = QApplication.instance()
            if app:
                app.setStyleSheet(TEMAS[nombre_tema])

            invalidar_cache_tema()

            self.repositorio_configuracion.guardar_preferencia(
                "tema",
                nombre_tema,
            )

    def _actualizar_estado_db(self, ruta: str | None = None):
        """
        Actualiza la etiqueta de la StatusBar con la DB activa.
        """

        if ruta:
            nombre = Path(ruta).name
            self.label_estado_db.setText(f"🟢 BD Activa: {nombre}")
        else:
            self.label_estado_db.setText("⚪ Sin base de datos cargada")

    def _configurar_ventana_y_tablas(self):
        """
        Configura el icono de la ventana y el ajuste de ancho de columnas
        en las tablas.
        """

        # Icono de ventana
        ruta_icono = Path("images/pine logo.png")
        if ruta_icono.exists():
            self.setWindowIcon(QIcon(str(ruta_icono)))

        # Ajuste de tablas para estirar columnas y ocupar todo el ancho disponible
        for tabla in (self.ui.tableViewCambios, self.ui.tableViewHistorico):
            tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

    def _actualizar_iconos_pestanas(self, indice_activo: int = 0):
        """
        Actualiza el color de los iconos de las pestañas:
        - Blanco (#FFFFFF) para la pestaña seleccionada
        - Azul/Acento (#106EBE) para las pestañas no seleccionadas
        """

        iconos = [
            "fa5s.balance-scale",
            "fa5s.layer-group",
            "fa5s.poll",
            "fa5s.history",
            "fa5s.question-circle",
        ]

        try:
            for i, nombre_icono in enumerate(iconos):
                if i >= self.ui.tabWidget.count():
                    break

                color = "#FFFFFF" if i == indice_activo else "#106EBE"
                self.ui.tabWidget.setTabIcon(i, qta.icon(nombre_icono, color=color))
        except Exception:
            pass

    def _configurar_iconos(self):
        """
        Asigna iconos a los elementos de la interfaz usando QtAwesome.
        """

        try:
            self._actualizar_iconos_pestanas(self.ui.tabWidget.currentIndex())

            # Iconos en blanco para los botones principales
            self.ui.pushButtonArchivo1.setIcon(qta.icon("fa5s.file-excel", color="#FFFFFF"))
            self.ui.pushButtonArchivo2.setIcon(qta.icon("fa5s.file-excel", color="#FFFFFF"))
            self.ui.pushButtonComparar.setIcon(qta.icon("fa5s.play", color="#FFFFFF"))
            self.ui.pushButtonGuardarComparacion.setIcon(qta.icon("fa5s.save", color="#FFFFFF"))
            self.ui.pushButtonAbrirHistorico.setIcon(qta.icon("fa5s.folder-open", color="#FFFFFF"))
            self.ui.pushButtonEliminarComparacion.setIcon(qta.icon("fa5s.trash-alt", color="#FFFFFF"))
            self.ui.pushButtonExportarHistorico.setIcon(qta.icon("fa5s.file-export", color="#FFFFFF"))

            # Botones de comparación múltiple
            self.ui.pushButtonAgregarArchivosMultiples.setIcon(qta.icon("fa5s.file-medical", color="#FFFFFF"))
            self.ui.pushButtonQuitarArchivoMultiple.setIcon(qta.icon("fa5s.minus-circle", color="#FFFFFF"))
            self.ui.pushButtonSubirArchivoMultiple.setIcon(qta.icon("fa5s.arrow-up", color="#FFFFFF"))
            self.ui.pushButtonBajarArchivoMultiple.setIcon(qta.icon("fa5s.arrow-down", color="#FFFFFF"))
            self.ui.pushButtonCompararMultiple.setIcon(qta.icon("fa5s.cogs", color="#FFFFFF"))

        except Exception:
            pass

    # =========================================================
    # EVENTOS
    # =========================================================

    def _conectar_eventos(self):
        """
        Conecta los controles de la interfaz con sus funciones.
        """

        self.ui.pushButtonArchivo1.clicked.connect(
            self.seleccionar_archivo_1
        )

        self.ui.pushButtonArchivo2.clicked.connect(
            self.seleccionar_archivo_nuevo
        )

        self.ui.comboBoxHoja1.currentTextChanged.connect(
            self.cargar_hoja_1
        )

        self.ui.comboBoxHoja2.currentTextChanged.connect(
            self.cargar_hoja_2
        )

        self.ui.pushButtonComparar.clicked.connect(
            self.comparar
        )

        self.ui.tableViewCambios.clicked.connect(
            self.mostrar_detalle_cambio
        )

        self.ui.tableViewHistorico.clicked.connect(
            self.mostrar_detalle_historico
        )

        self.ui.pushButtonGuardarComparacion.clicked.connect(
            self.guardar_comparacion
        )

        self.ui.pushButtonAbrirHistorico.clicked.connect(
            self.abrir_historico
        )

        self.ui.pushButtonEliminarComparacion.clicked.connect(
            self.eliminar_comparacion_historico
        )

        self.ui.lineEditBuscarClaveHistorico.textChanged.connect(
            lambda _: self.filtrar_historico()
        )

        self.ui.comboBoxIdentificadorHistorico.currentTextChanged.connect(
            lambda _: self.filtrar_historico()
        )

        self.ui.pushButtonExportarHistorico.clicked.connect(
            self.exportar_historico_excel
        )

        # Eventos comparación múltiple
        self.ui.pushButtonAgregarArchivosMultiples.clicked.connect(
            self.agregar_archivos_multiples
        )

        self.ui.pushButtonQuitarArchivoMultiple.clicked.connect(
            self.quitar_archivo_multiple
        )

        self.ui.pushButtonSubirArchivoMultiple.clicked.connect(
            self.subir_archivo_multiple
        )

        self.ui.pushButtonBajarArchivoMultiple.clicked.connect(
            self.bajar_archivo_multiple
        )

        self.ui.comboBoxHojaMultiple.currentTextChanged.connect(
            self.cargar_hoja_multiple
        )

        self.ui.lineEditBuscarClaveMultiple.textChanged.connect(
            lambda texto: self._filtrar_lista(
                self.ui.listaColumnasClaveMultiple,
                texto,
            )
        )

        self.ui.lineEditBuscarCompararMultiple.textChanged.connect(
            lambda texto: self._filtrar_lista(
                self.ui.listaColumnasCompararMultiple,
                texto,
            )
        )

        self.ui.pushButtonCompararMultiple.clicked.connect(
            self.procesar_comparacion_multiple
        )

    def _filtrar_lista(
        self,
        lista,
        texto: str,
    ):
        texto = texto.strip().lower()

        for i in range(lista.count()):

            item = lista.item(i)

            visible = (
                texto == ""
                or texto in item.text().lower()
            )

            item.setHidden(
                not visible
            )

    def _texto_con_diferencias(
        self,
        texto_anterior: str,
        texto_nuevo: str,
        mostrar_anterior: bool,
    ) -> str:
        """
        Genera HTML resaltando con fondo amarillo suave las partes
        diferentes entre el valor anterior y el nuevo.

        El color se adapta al tema claro/oscuro de Qt.
        """

        # ---------------------------------------------------------
        # Color de resaltado según el tema
        # ---------------------------------------------------------

        from PySide6.QtWidgets import QApplication
        from PySide6.QtGui import QPalette

        paleta = QApplication.palette()

        fondo = paleta.color(
            QPalette.ColorRole.Base
        )

        # Amarillo de referencia
        amarillo = (220, 180, 40)

        # En función del fondo actual hacemos una mezcla
        porcentaje = 0.35

        rojo = int(
            fondo.red() * (1 - porcentaje)
            + amarillo[0] * porcentaje
        )

        verde = int(
            fondo.green() * (1 - porcentaje)
            + amarillo[1] * porcentaje
        )

        azul = int(
            fondo.blue() * (1 - porcentaje)
            + amarillo[2] * porcentaje
        )

        color_resaltado = (
            f"rgb({rojo},{verde},{azul})"
        )

        # ---------------------------------------------------------
        # Comparar textos
        # ---------------------------------------------------------

        matcher = SequenceMatcher(
            None,
            texto_anterior,
            texto_nuevo,
        )

        resultado = []

        for etiqueta, i1, i2, j1, j2 in matcher.get_opcodes():

            if mostrar_anterior:
                texto = texto_anterior[i1:i2]
            else:
                texto = texto_nuevo[j1:j2]

            if not texto:
                continue

            # -----------------------------------------------------
            # Escapar HTML
            # -----------------------------------------------------

            texto = (
                texto
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace("\n", "<br>")
            )

            # -----------------------------------------------------
            # Resaltar diferencia
            # -----------------------------------------------------

            if etiqueta != "equal":

                texto = (
                    f'<span style="background-color: {color_resaltado};">'
                    f"{texto}"
                    "</span>"
                )

            resultado.append(texto)

        return "".join(resultado)

    def _mostrar_valores_con_diferencias(
        self,
        valor_1,
        valor_2,
        edit1,
        edit2
    ):
        """
        Muestra los valores anterior y nuevo resaltando
        las diferencias.
        """

        texto_anterior = (
            ""
            if valor_1 is None
            else str(valor_1)
        )

        texto_nuevo = (
            ""
            if valor_2 is None
            else str(valor_2)
        )

        # Si ambos valores son iguales, no hay nada que resaltar.
        if texto_anterior == texto_nuevo:
            edit1.setPlainText(texto_anterior)
            edit2.setPlainText(texto_nuevo)

            return

        html_anterior = self._texto_con_diferencias(
            texto_anterior,
            texto_nuevo,
            True,
        )

        html_nuevo = self._texto_con_diferencias(
            texto_anterior,
            texto_nuevo,
            True,
        )

        edit1.setHtml(html_anterior)
        edit2.setHtml(html_nuevo)

    def _crear_tabla_filtrable(
        self,
        nombre_tabla,
    ):
        """
        Sustituye un QTableView por QFilterableTableView.
        """

        tabla_original = getattr(
            self.ui,
            nombre_tabla,
        )

        tabla_filtrable = QFilterableTableView(
            tabla_original.parentWidget()
        )

        tabla_filtrable.setObjectName(
            nombre_tabla
        )

        # ---------------------------------------------------------
        # Copiar propiedades visuales
        # ---------------------------------------------------------

        tabla_filtrable.setMinimumSize(
            tabla_original.minimumSize()
        )

        tabla_filtrable.setSizePolicy(
            tabla_original.sizePolicy()
        )

        tabla_filtrable.setAlternatingRowColors(
            tabla_original.alternatingRowColors()
        )

        tabla_filtrable.setSortingEnabled(
            True
        )

        # ---------------------------------------------------------
        # Sustituir dentro del layout
        # ---------------------------------------------------------

        layout = tabla_original.parentWidget().layout()

        if layout is None:
            raise RuntimeError(
                f"No se ha encontrado el layout de {nombre_tabla}"
            )

        layout.replaceWidget(
            tabla_original,
            tabla_filtrable,
        )

        tabla_original.hide()
        tabla_original.deleteLater()

        # ---------------------------------------------------------
        # Actualizar referencia de Ui_MainWindow
        # ---------------------------------------------------------

        setattr(
            self.ui,
            nombre_tabla,
            tabla_filtrable,
        )

    def seleccionar_archivo_1(self):
        """
        Permite seleccionar el Excel 1.
        """

        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Excel 1",
            "",
            "Archivos Excel (*.xlsx *.xls)",
        )

        if not ruta:
            return

        self.ui.lineEditArchivo1.setText(ruta)

        self.df_1 = None

        self._cargar_hojas(
            ruta,
            self.ui.comboBoxHoja1,
        )

        self._actualizar_columnas()

    # =========================================================
    # ARCHIVO NUEVO
    # =========================================================

    def seleccionar_archivo_nuevo(self):
        """
        Permite seleccionar el Excel nuevo.
        """

        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Excel nuevo",
            "",
            "Archivos Excel (*.xlsx *.xls)",
        )

        if not ruta:
            return

        self.ui.lineEditArchivo2.setText(ruta)

        self.df_2 = None

        self._cargar_hojas(
            ruta,
            self.ui.comboBoxHoja2,
        )

        self._actualizar_columnas()

    # =========================================================
    # CARGAR HOJAS
    # =========================================================

    def _cargar_hojas(
        self,
        ruta: str,
        combo_box,
    ):
        """
        Carga las hojas del Excel seleccionado.
        """

        try:

            hojas = obtener_hojas(ruta)

            combo_box.blockSignals(True)

            combo_box.clear()
            combo_box.addItems(hojas)

            combo_box.blockSignals(False)

            # Si hay hojas, cargamos la primera
            if hojas:
                self._cargar_hoja_seleccionada(
                    combo_box
                )

        except Exception as error:

            combo_box.blockSignals(False)

            QMessageBox.critical(
                self,
                "Error al leer Excel",
                f"No se ha podido leer el archivo:\n\n{error}",
            )

    # =========================================================
    # HOJA 1
    # =========================================================

    def cargar_hoja_1(
        self,
        nombre_hoja: str,
    ):
        """
        Carga la hoja seleccionada del Excel 1.
        """

        if not nombre_hoja:
            self.df_1 = None
            self._actualizar_columnas()
            return

        ruta = self.ui.lineEditArchivo1.text()

        if not ruta:
            return

        try:

            self.df_1 = leer_excel(
                ruta,
                nombre_hoja,
            )

            self._actualizar_columnas()

        except Exception as error:

            self.df_1 = None

            QMessageBox.critical(
                self,
                "Error al leer hoja",
                f"No se ha podido leer la hoja:\n\n{error}",
            )

    # =========================================================
    # HOJA 2
    # =========================================================

    def cargar_hoja_2(
        self,
        nombre_hoja: str,
    ):
        """
        Carga la hoja seleccionada del Excel nuevo.
        """

        if not nombre_hoja:
            self.df_2 = None
            self._actualizar_columnas()
            return

        ruta = self.ui.lineEditArchivo2.text()

        if not ruta:
            return

        try:

            self.df_2 = leer_excel(
                ruta,
                nombre_hoja,
            )

            self._actualizar_columnas()

        except Exception as error:

            self.df_2 = None

            QMessageBox.critical(
                self,
                "Error al leer hoja",
                f"No se ha podido leer la hoja:\n\n{error}",
            )

    # =========================================================
    # CARGAR HOJA SELECCIONADA
    # =========================================================

    def _cargar_hoja_seleccionada(
        self,
        combo_box,
    ):
        """
        Fuerza la carga de la hoja actualmente seleccionada.
        """

        nombre_hoja = combo_box.currentText()

        if combo_box is self.ui.comboBoxHoja1:
            self.cargar_hoja_1(nombre_hoja)

        elif combo_box is self.ui.comboBoxHoja2:
            self.cargar_hoja_2(nombre_hoja)

    # =========================================================
    # ACTUALIZAR COLUMNAS
    # =========================================================

    def _actualizar_columnas(self):
        """
        Actualiza las columnas disponibles para seleccionar.

        Sólo se muestran las columnas presentes en ambos Excel.
        """

        # Todavía no tenemos ambos DataFrames
        if (
            self.df_1 is None
            or self.df_2 is None
        ):
            self.ui.listaColumnasClave.clear()
            self.ui.listaColumnasComparar.clear()
            return
        
        self.firma_configuracion_actual = (
            generar_firma_excel(
                self.df_1,
                self.df_2,
            )
        )
            
        columnas_comunes = obtener_columnas_comunes(
            self.df_1,
            self.df_2,
        )

        columnas_2 = obtener_columnas_nuevas(
            self.df_1,
            self.df_2,
        )

        columnas_eliminadas = obtener_columnas_eliminadas(
            self.df_1,
            self.df_2,
        )

        self.ui.listaColumnasClave.clear()
        self.ui.listaColumnasComparar.clear()

        self.ui.listaColumnasClave.addItems(
            columnas_comunes
        )

        self.ui.listaColumnasComparar.addItems(
            columnas_comunes
        )

        # Mostrar información de estructura
        self._mostrar_informacion_columnas(
            columnas_2,
            columnas_eliminadas,
        )

        self._cargar_configuracion_cacheada()

    # =========================================================
    # INFORMACIÓN DE COLUMNAS
    # =========================================================

    def _mostrar_informacion_columnas(
        self,
        columnas_2: list[str],
        columnas_eliminadas: list[str],
    ):
        """
        Muestra información sobre diferencias estructurales.
        """

        mensajes = []

        if columnas_2:

            mensajes.append(
                "Columnas 2: "
                + ", ".join(columnas_2)
            )

        if columnas_eliminadas:

            mensajes.append(
                "Columnas eliminadas: "
                + ", ".join(columnas_eliminadas)
            )

        if mensajes:

            self.ui.statusBar.showMessage(
                " | ".join(mensajes)
            )

        else:

            self.ui.statusBar.showMessage(
                "Las estructuras de ambos Excel coinciden."
            )

    def _cargar_configuracion_cacheada(self):
        """
        Busca una configuración guardada para la estructura
        actual de los dos Excel y la aplica si existe.
        """

        if not self.firma_configuracion_actual:
            return False

        configuracion = (
            self.repositorio_configuracion
            .obtener_configuracion(
                self.firma_configuracion_actual
            )
        )

        if configuracion is None:
            return False

        # ---------------------------------------------------------
        # Seleccionar columnas clave
        # ---------------------------------------------------------

        self._seleccionar_items_lista(
            self.ui.listaColumnasClave,
            configuracion["columnas_clave"],
        )

        # ---------------------------------------------------------
        # Seleccionar columnas a comparar
        # ---------------------------------------------------------

        self._seleccionar_items_lista(
            self.ui.listaColumnasComparar,
            configuracion["columnas_comparadas"],
        )

        self.ui.statusBar.showMessage(
            "Configuración anterior recuperada automáticamente."
        )

        return True

    def _seleccionar_items_lista(
        self,
        lista,
        valores: list[str],
    ):
        """
        Selecciona en una QListWidget los elementos indicados.
        """

        valores = set(valores)

        for i in range(lista.count()):

            item = lista.item(i)

            item.setSelected(
                item.text() in valores
            )

    # =========================================================
    # COMPARAR
    # =========================================================

    def comparar(self):
        """
        Ejecuta la comparación de los dos Excel.
        """

        # ---------------------------------------------------------
        # Comprobar que tenemos los dos archivos
        # ---------------------------------------------------------

        if self.df_1 is None:

            QMessageBox.warning(
                self,
                "Falta el Excel anterior",
                "Selecciona el Excel anterior.",
            )

            return

        if self.df_2 is None:

            QMessageBox.warning(
                self,
                "Falta el Excel nuevo",
                "Selecciona el Excel nuevo.",
            )

            return

        # ---------------------------------------------------------
        # Obtener columnas seleccionadas
        # ---------------------------------------------------------

        columnas_clave = [
            item.text()
            for item in self.ui.listaColumnasClave.selectedItems()
        ]

        columnas_comparar = [
            item.text()
            for item in self.ui.listaColumnasComparar.selectedItems()
        ]

        self.repositorio_configuracion.guardar_configuracion(
            firma=self.firma_configuracion_actual,
            hoja_anterior=self.ui.comboBoxHoja1.currentText(),
            hoja_nueva=self.ui.comboBoxHoja2.currentText(),
            columnas_clave=columnas_clave,
            columnas_comparadas=columnas_comparar,
        )

        # ---------------------------------------------------------
        # Validar clave
        # ---------------------------------------------------------

        if not columnas_clave:

            QMessageBox.warning(
                self,
                "Falta la clave",
                "Selecciona al menos una columna clave.",
            )

            return

        # ---------------------------------------------------------
        # Validar columnas a comparar
        # ---------------------------------------------------------

        if not columnas_comparar:

            QMessageBox.warning(
                self,
                "Faltan columnas",
                "Selecciona al menos una columna a comparar.",
            )

            return

        # ---------------------------------------------------------
        # Ejecutar comparación en segundo plano con diálogo de progreso
        # ---------------------------------------------------------

        from src.excel.comparador import comparar_dataframes

        dialogo = QProgressDialog("Comparando archivos...", None, 0, 0, self)
        dialogo.setWindowTitle("Procesando")
        dialogo.setCancelButton(None)
        dialogo.setModal(True)
        dialogo.show()

        class WorkerComparar(QThread):
            terminado = Signal(object, object)

            def __init__(self, df1, df2, claves, comparar):
                super().__init__()
                self.df1 = df1
                self.df2 = df2
                self.claves = claves
                self.comparar = comparar

            def run(self):
                try:
                    res = comparar_dataframes(
                        self.df1,
                        self.df2,
                        columnas_clave=self.claves,
                        columnas_comparar=self.comparar,
                    )
                    self.terminado.emit(res, None)
                except Exception as ex:
                    self.terminado.emit(None, ex)

        self.worker = WorkerComparar(self.df_1, self.df_2, columnas_clave, columnas_comparar)

        def _al_terminar(cambios, error):
            dialogo.close()
            if error:
                QMessageBox.critical(
                    self,
                    "Error durante la comparación",
                    f"No se ha podido comparar los archivos:\n\n{error}",
                )
                return

            self.modelo_cambios.actualizar(cambios)
            self.ui.tabWidget.setCurrentWidget(self.ui.tab_resultados)
            self.ui.labelResumen.setText(f"{len(cambios)} cambios encontrados")
            self.ui.tableViewCambios.resizeColumnsToContents()
            self.ui.statusBar.showMessage("Comparación completada correctamente.")

        self.worker.terminado.connect(_al_terminar)
        self.worker.start()

    def _obtener_fila_modelo_origen(self, index):
        """
        Obtiene la fila correspondiente al ModeloCambios
        partiendo del índice de la tabla filtrable.
        """

        indice = index

        while indice.isValid():

            modelo = indice.model()

            if modelo is self.modelo_cambios:
                return indice.row()

            if modelo is self.modelo_historico:
                return indice.row()

            if hasattr(modelo, "mapToSource"):

                indice = modelo.mapToSource(indice)

            else:
                break

        return -1

    def mostrar_detalle_cambio(self, index):
        
        if not index.isValid():
            return

        fila = self._obtener_fila_modelo_origen(index)

        if fila < 0:
            return

        cambio = self.modelo_cambios.obtener_cambio(fila)

        if cambio is None:
            return

        self._mostrar_valores_con_diferencias(
            cambio.valor_1,
            cambio.valor_2,
            self.ui.textEditValor1,
            self.ui.textEditValor2
        )

    def guardar_comparacion(self):
        """
        Guarda la comparación actual en una base de datos SQLite.

        El usuario puede:
        - Crear una nueva base de datos.
        - Añadir la comparación a una base de datos existente.
        """

        # =========================================================
        # COMPROBAR QUE EXISTE UNA COMPARACIÓN
        # =========================================================

        if not self.modelo_cambios.cambios:

            QMessageBox.warning(
                self,
                "Sin comparación",
                "Primero debes realizar una comparación.",
            )

            return

        # =========================================================
        # OBTENER ARCHIVOS ACTUALES
        # =========================================================

        archivo_anterior = self.ui.lineEditArchivo1.text().strip()
        archivo_nuevo = self.ui.lineEditArchivo2.text().strip()

        if not archivo_anterior or not archivo_nuevo:

            QMessageBox.warning(
                self,
                "Archivos no seleccionados",
                "Debes seleccionar los dos archivos Excel antes de guardar.",
            )

            return

        # =========================================================
        # ELEGIR DESTINO
        # =========================================================

        respuesta = QMessageBox.question(
            self,
            "Guardar comparación",
            "¿Quieres crear una nueva base de datos?\n\n"
            "Sí → Crear una nueva base de datos\n"
            "No → Utilizar una base de datos existente",
            buttons=(
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No
                | QMessageBox.StandardButton.Cancel
            ),
            defaultButton=QMessageBox.StandardButton.Yes,
        )

        if respuesta == QMessageBox.StandardButton.Cancel:
            return

        # =========================================================
        # CREAR NUEVA SQLITE
        # =========================================================

        if respuesta == QMessageBox.StandardButton.Yes:

            nombre_bd = obtener_nombre_bd(
                archivo_anterior,
                archivo_nuevo,
            )

            ruta, _ = QFileDialog.getSaveFileName(
                self,
                "Crear base de datos SQLite",
                nombre_bd,
                "Base de datos SQLite (*.sqlite *.db)",
            )

            if not ruta:
                return

        # =========================================================
        # UTILIZAR SQLITE EXISTENTE
        # =========================================================

        else:

            ruta, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar base de datos SQLite",
                "",
                "Base de datos SQLite (*.sqlite *.db)",
            )

            if not ruta:
                return

        # =========================================================
        # PROPONER IDENTIFICADOR
        # =========================================================

        identificador_propuesto = obtener_identificador_propuesto(
            archivo_anterior,
            archivo_nuevo,
        )

        identificador, aceptado = QInputDialog.getText(
            self,
            "Identificador de comparación",
            "Introduce el identificador de esta comparación:",
            text=identificador_propuesto,
        )

        if not aceptado:
            return

        identificador = identificador.strip()

        if not identificador:

            QMessageBox.warning(
                self,
                "Identificador vacío",
                "Debes introducir un identificador.",
            )

            return

        # =========================================================
        # OBTENER CONFIGURACIÓN ACTUAL
        # =========================================================

        columnas_clave = [
            item.text()
            for item in self.ui.listaColumnasClave.selectedItems()
        ]

        columnas_comparar = [
            item.text()
            for item in self.ui.listaColumnasComparar.selectedItems()
        ]

        hoja_anterior = self.ui.comboBoxHoja1.currentText()
        hoja_nueva = self.ui.comboBoxHoja2.currentText()

        # =========================================================
        # GUARDAR
        # =========================================================

        try:

            repositorio = RepositorioSQLite(ruta)

            comparacion_id = repositorio.guardar_comparacion(
                identificador=identificador,
                archivo_anterior=archivo_anterior,
                archivo_nuevo=archivo_nuevo,
                hoja_anterior=hoja_anterior,
                hoja_nueva=hoja_nueva,
                columnas_clave=columnas_clave,
                columnas_comparadas=columnas_comparar,
                cambios=self.modelo_cambios.cambios,
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "Error al guardar",
                "No se ha podido guardar la comparación:\n\n"
                f"{error}",
            )

            return

        # =========================================================
        # CONFIRMACIÓN
        # =========================================================

        QMessageBox.information(
            self,
            "Comparación guardada",
            "La comparación se ha guardado correctamente.\n\n"
            f"Identificador: {identificador}\n"
            f"Cambios guardados: "
            f"{len(self.modelo_cambios.cambios)}\n\n"
            f"Base de datos:\n{ruta}",
        )

        self.ui.statusBar.showMessage(
            f"Comparación guardada: {identificador}"
        )

        self._actualizar_estado_db(ruta)

    def _actualizar_identificadores_historico(self, cambios):
        combo = self.ui.comboBoxIdentificadorHistorico

        combo.blockSignals(True)

        combo.clear()
        combo.addItem("Todos")

        identificadores = sorted(
            {
                str(cambio["identificador"])
                for cambio in cambios
                if cambio["identificador"] is not None
            }
        )

        combo.addItems(identificadores)

        combo.blockSignals(False)

    def _obtener_modelo_visible_historico(self):
        """
        Obtiene el modelo que está utilizando actualmente
        la tabla del histórico.

        Si QFilterableTableView utiliza un modelo proxy
        para filtrar/ordenar, devuelve dicho modelo.
        """

        modelo = self.ui.tableViewHistorico.model()

        if modelo is None:
            return None

        return modelo

    def abrir_historico(self):

        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar base de datos SQLite",
            "",
            "Base de datos SQLite (*.sqlite *.db)",
        )

        if not ruta:
            return

        try:

            from src.database.repositorio import (
                RepositorioSQLite,
            )

            repositorio = RepositorioSQLite(ruta)

            cambios = (
                repositorio.obtener_historico_cambios()
            )

            self.ruta_historico = ruta
            self.repositorio_historico = repositorio

            self._actualizar_identificadores_historico(
                cambios
            )
                        
            self.modelo_historico.actualizar(
                cambios
            )

            self.limpiar_detalle_historico()

            self.ui.lineEditBaseHistorico.setText(
                ruta
            )

            self.ui.tableViewHistorico.resizeColumnsToContents()

            self.ui.statusBar.showMessage(
                "Histórico cargado correctamente."
            )

            self._actualizar_estado_db(ruta)

        except Exception as error:

            QMessageBox.critical(
                self,
                "Error",
                f"No se ha podido abrir la base de datos:\n\n{error}",
            )

    def exportar_historico_excel(self):
        """
        Exporta a Excel exactamente los registros que se están
        mostrando actualmente en la tabla del histórico.

        Respeta:
        - filtros aplicados
        - orden de las filas
        - columnas visibles
        """

        tabla = self.ui.tableViewHistorico

        # =========================================================
        # OBTENER EL PROXY REAL DE FILTRADO
        # =========================================================

        modelo = tabla._filter_proxy

        if modelo is None:

            QMessageBox.warning(
                self,
                "Sin datos",
                "No hay datos para exportar.",
            )

            return

        # =========================================================
        # COMPROBAR FILAS
        # =========================================================

        numero_filas = modelo.rowCount()

        if numero_filas == 0:

            QMessageBox.warning(
                self,
                "Sin datos",
                "No hay registros visibles para exportar.",
            )

            return

        # =========================================================
        # ELEGIR ARCHIVO
        # =========================================================

        nombre_sugerido = "Historico.xlsx"
        if hasattr(self, "ruta_historico") and self.ruta_historico:
            stem = Path(self.ruta_historico).stem
            if stem:
                nombre_sugerido = f"{stem}.xlsx"

        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar histórico a Excel",
            nombre_sugerido,
            "Archivos Excel (*.xlsx)",
        )

        if not ruta:
            return

        try:

            libro = Workbook()

            hoja = libro.active

            hoja.title = "Histórico"

            # =====================================================
            # COLUMNAS
            # =====================================================

            columnas = modelo.columnCount()

            columnas_exportar = []

            for columna in range(columnas):

                # Comprobar si la columna está visible
                if tabla.isColumnHidden(columna):
                    continue

                encabezado = modelo.headerData(
                    columna,
                    Qt.Orientation.Horizontal,
                    Qt.ItemDataRole.DisplayRole,
                )

                columnas_exportar.append(
                    (
                        columna,
                        encabezado,
                    )
                )

            # =====================================================
            # CABECERAS
            # =====================================================

            for numero_columna, (_, encabezado) in enumerate(
                columnas_exportar,
                start=1,
            ):

                celda = hoja.cell(
                    row=1,
                    column=numero_columna,
                    value=encabezado,
                )

                celda.font = Font(
                    bold=True
                )

            # =====================================================
            # FILAS FILTRADAS
            # =====================================================

            for fila in range(numero_filas):

                for numero_columna, (
                    columna,
                    _,
                ) in enumerate(
                    columnas_exportar,
                    start=1,
                ):

                    indice = modelo.index(
                        fila,
                        columna,
                    )

                    valor = modelo.data(
                        indice,
                        Qt.ItemDataRole.DisplayRole,
                    )

                    hoja.cell(
                        row=fila + 2,
                        column=numero_columna,
                        value=valor,
                    )

            # =====================================================
            # FORMATO
            # =====================================================

            hoja.freeze_panes = "A2"

            hoja.auto_filter.ref = hoja.dimensions

            # Ajustar ancho de columnas
            for columna in hoja.columns:

                longitud = 0

                for celda in columna:

                    if celda.value is not None:

                        longitud = max(
                            longitud,
                            len(str(celda.value)),
                        )

                hoja.column_dimensions[
                    columna[0].column_letter
                ].width = min(
                    longitud + 2,
                    50,
                )

            # =====================================================
            # GUARDAR
            # =====================================================

            libro.save(ruta)

            # =====================================================
            # CONFIRMACIÓN
            # =====================================================

            QMessageBox.information(
                self,
                "Exportación completada",
                (
                    "El histórico se ha exportado correctamente.\n\n"
                    f"Registros exportados: {numero_filas}\n"
                    f"Archivo:\n{ruta}"
                ),
            )

            self.ui.statusBar.showMessage(
                f"Histórico exportado: "
                f"{numero_filas} registros."
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "Error al exportar",
                (
                    "No se ha podido exportar el histórico:\n\n"
                    f"{error}"
                ),
            )
        
    def filtrar_historico(self):

        if self.repositorio_historico is None:
            return

        texto_clave = (
            self.ui.lineEditBuscarClaveHistorico
            .text()
            .strip()
            .lower()
        )

        identificador_seleccionado = (
            self.ui.comboBoxIdentificadorHistorico
            .currentText()
            .strip()
        )

        cambios = (
            self.repositorio_historico
            .obtener_historico_cambios()
        )

        cambios_filtrados = []

        for cambio in cambios:

            clave = str(
                cambio["clave"]
            ).lower()

            identificador = str(
                cambio["identificador"]
            )

            coincide_clave = (
                not texto_clave
                or texto_clave in clave
            )

            if identificador_seleccionado == "Todos":
                coincide_identificador = True
            else:
                coincide_identificador = (
                    identificador
                    == identificador_seleccionado
                )

            if (
                coincide_clave
                and coincide_identificador
            ):
                cambios_filtrados.append(cambio)

        self.modelo_historico.actualizar(
            cambios_filtrados
        )

        self.limpiar_detalle_historico()

    def eliminar_comparacion_historico(self):

        index = (
            self.ui.tableViewHistorico
            .currentIndex()
        )

        if not index.isValid():

            QMessageBox.warning(
                self,
                "Sin selección",
                "Selecciona una comparación.",
            )

            return

        cambio = (
            self.modelo_historico
            .obtener_cambio(index.row())
        )

        if cambio is None:
            return

        comparacion_id = cambio["comparacion_id"]

        identificador = cambio["identificador"]
        if comparacion_id is None:

            QMessageBox.critical(
                self,
                "Error",
                "No se ha podido identificar la comparación.",
            )

            return

        respuesta = QMessageBox.question(
            self,
            "Eliminar comparación",
            (
                "¿Deseas eliminar la comparación completa?\n\n"
                f"Identificador: {identificador}\n\n"
                "Se eliminarán todos los cambios asociados "
                "a esta comparación."
            ),
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if respuesta != QMessageBox.StandardButton.Yes:
            return

        try:

            self.repositorio_historico.eliminar_comparacion(
                comparacion_id
            )

            cambios = (
                self.repositorio_historico
                .obtener_historico_cambios()
            )

            self.modelo_historico.actualizar(
                cambios
            )

            self.limpiar_detalle_historico()

            self.ui.statusBar.showMessage(
                f"Comparación '{identificador}' eliminada."
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "Error",
                (
                    "No se ha podido eliminar la comparación:\n\n"
                    f"{error}"
                ),
            )

    def mostrar_detalle_historico(self, index):
        if not index.isValid():
            return

        fila = self._obtener_fila_modelo_origen(index)

        if fila < 0:
            return

        cambio = self.modelo_historico.obtener_cambio(fila)

        if cambio is None:
            return

        self._mostrar_valores_con_diferencias(
            cambio["valor_1"],
            cambio["valor_2"],
            self.ui.textEditValor1Historico,
            self.ui.textEditValor2Historico
        )

        self.ui.labelResumenHistorico.setText(
            cambio["fecha"]
        )

    def limpiar_detalle_historico(self):

        self.ui.labelResumenHistorico.setText(
            "Selecciona un cambio"
        )

        self.ui.textEditValor1Historico.clear()
        self.ui.textEditValor2Historico.clear()

    # =========================================================
    # LÓGICA DE COMPARACIÓN MÚLTIPLE / LOTE
    # =========================================================

    def agregar_archivos_multiples(self):
        """
        Añade múltiples archivos Excel a la lista secuencial.
        """

        rutas, _ = QFileDialog.getOpenFileNames(
            self,
            "Seleccionar archivos Excel para comparación múltiple",
            "",
            "Archivos Excel (*.xlsx *.xls)",
        )

        if not rutas:
            return

        lista = self.ui.listaArchivosMultiples

        # Obtener rutas ya existentes
        rutas_existentes = {
            lista.item(i).text() for i in range(lista.count())
        }

        nuevas_rutas = [r for r in rutas if r not in rutas_existentes]

        # Ordenar alfabéticamente las nuevas rutas
        nuevas_rutas.sort()

        for ruta in nuevas_rutas:
            lista.addItem(ruta)

        self._actualizar_hojas_multiples()

    def quitar_archivo_multiple(self):
        """
        Elimina el archivo seleccionado de la lista múltiple.
        """

        lista = self.ui.listaArchivosMultiples
        fila = lista.currentRow()

        if fila >= 0:
            lista.takeItem(fila)
            self._actualizar_hojas_multiples()

    def subir_archivo_multiple(self):
        """
        Sube una posición el archivo seleccionado.
        """

        lista = self.ui.listaArchivosMultiples
        fila = lista.currentRow()

        if fila > 0:
            item = lista.takeItem(fila)
            lista.insertItem(fila - 1, item)
            lista.setCurrentRow(fila - 1)

    def bajar_archivo_multiple(self):
        """
        Baja una posición el archivo seleccionado.
        """

        lista = self.ui.listaArchivosMultiples
        fila = lista.currentRow()

        if 0 <= fila < lista.count() - 1:
            item = lista.takeItem(fila)
            lista.insertItem(fila + 1, item)
            lista.setCurrentRow(fila + 1)

    def _obtener_rutas_multiples(self) -> list[str]:
        """
        Devuelve la lista de rutas ordenadas según la UI.
        """
        lista = self.ui.listaArchivosMultiples
        return [lista.item(i).text() for i in range(lista.count())]

    def _actualizar_hojas_multiples(self):
        """
        Obtiene las hojas comunes presentes en TODOS los Excel seleccionados.
        """

        lista = self.ui.listaArchivosMultiples
        combo = self.ui.comboBoxHojaMultiple

        combo.blockSignals(True)
        combo.clear()

        rutas = [lista.item(i).text() for i in range(lista.count())]

        if not rutas:
            combo.blockSignals(False)

            self.ui.listaColumnasClaveMultiple.clear()
            self.ui.listaColumnasCompararMultiple.clear()

            return

        try:
            hojas_comunes = None

            for ruta in rutas:
                hojas = set(obtener_hojas(ruta))
                if hojas_comunes is None:
                    hojas_comunes = hojas
                else:
                    hojas_comunes &= hojas

            lista_hojas = sorted(list(hojas_comunes or []))
            combo.addItems(lista_hojas)
            combo.blockSignals(False)

            if lista_hojas:
                self.cargar_hoja_multiple(combo.currentText())
            else:
                self.ui.listaColumnasClaveMultiple.clear()
                self.ui.listaColumnasCompararMultiple.clear()

        except Exception as error:
            combo.blockSignals(False)
            QMessageBox.critical(
                self,
                "Error al leer hojas comunes",
                f"Ocurrió un error al inspeccionar las hojas:\n\n{error}",
            )

    def cargar_hoja_multiple(self, nombre_hoja: str):
        """
        Carga las columnas comunes a todos los Excel para la hoja dada.
        """

        lista = self.ui.listaArchivosMultiples
        rutas = [lista.item(i).text() for i in range(lista.count())]

        if not rutas or not nombre_hoja:
            self.ui.listaColumnasClaveMultiple.clear()
            self.ui.listaColumnasCompararMultiple.clear()

            return

        try:
            columnas_comunes = None

            for ruta in rutas:
                df = leer_excel(ruta, nombre_hoja)
                cols = set(df.columns)
                if columnas_comunes is None:
                    columnas_comunes = cols
                else:
                    columnas_comunes &= cols

            columnas = sorted(list(columnas_comunes or []))

            self.ui.listaColumnasClaveMultiple.clear()
            self.ui.listaColumnasCompararMultiple.clear()

            self.ui.listaColumnasClaveMultiple.addItems(columnas)
            self.ui.listaColumnasCompararMultiple.addItems(columnas)

        except Exception as error:
            QMessageBox.critical(
                self,
                "Error al leer columnas",
                f"No se pudieron leer las columnas de los archivos:\n\n{error}",
            )

    def procesar_comparacion_multiple(self):
        """
        Ejecuta la comparación secuencial de N archivos (1 → 2 → 3 ...)
        y guarda los resultados directamente en una base de datos SQLite.
        """

        lista = self.ui.listaArchivosMultiples
        rutas = [lista.item(i).text() for i in range(lista.count())]

        if len(rutas) < 2:
            QMessageBox.warning(
                self,
                "Faltan archivos",
                "Debes agregar al menos 2 archivos Excel para comparar.",
            )

            return

        nombre_hoja = self.ui.comboBoxHojaMultiple.currentText()

        if not nombre_hoja:
            QMessageBox.warning(
                self,
                "Falta la hoja",
                "Selecciona una hoja válida para comparar.",
            )

            return

        columnas_clave = [
            item.text()
            for item in self.ui.listaColumnasClaveMultiple.selectedItems()
        ]

        columnas_comparar = [
            item.text()
            for item in self.ui.listaColumnasCompararMultiple.selectedItems()
        ]

        if not columnas_clave:
            QMessageBox.warning(
                self,
                "Falta la clave",
                "Selecciona al menos una columna clave.",
            )

            return

        if not columnas_comparar:
            QMessageBox.warning(
                self,
                "Faltan columnas",
                "Selecciona al menos una columna a comparar.",
            )

            return

        # Pedir destino SQLite
        nombre_bd = obtener_nombre_bd(rutas[0], rutas[-1])
        ruta_bd, _ = QFileDialog.getSaveFileName(
            self,
            "Crear base de datos SQLite para Histórico Múltiple",
            nombre_bd,
            "Base de datos SQLite (*.sqlite *.db)",
        )

        if not ruta_bd:
            return

        from src.excel.comparador import comparar_dataframes

        pasos_totales = len(rutas) - 1
        dialogo = QProgressDialog("Procesando archivos...", "Cancelar", 0, pasos_totales, self)
        dialogo.setWindowTitle("Procesando Comparación Múltiple")
        dialogo.setModal(True)
        dialogo.show()

        class WorkerMultiple(QThread):
            progreso = Signal(int, str)
            finalizado = Signal(int, object)

            def __init__(self, rutas_files, hoja, claves, comparar, db_path):
                super().__init__()
                self.rutas = rutas_files
                self.hoja = hoja
                self.claves = claves
                self.comparar = comparar
                self.db_path = db_path

            def run(self):
                try:
                    repositorio = RepositorioSQLite(self.db_path)
                    total_c = 0

                    df_ant = leer_excel(self.rutas[0], self.hoja)

                    for i in range(len(self.rutas) - 1):
                        if self.isInterrupted():
                            break

                        r_ant = self.rutas[i]
                        r_nue = self.rutas[i + 1]

                        self.progreso.emit(i, f"Comparando ({i + 1}/{len(self.rutas) - 1}): {Path(r_nue).name}")

                        df_nue = leer_excel(r_nue, self.hoja)
                        ident = obtener_identificador_propuesto(r_ant, r_nue)

                        cambios = comparar_dataframes(
                            df_ant,
                            df_nue,
                            columnas_clave=self.claves,
                            columnas_comparar=self.comparar,
                        )

                        repositorio.guardar_comparacion(
                            identificador=ident,
                            archivo_anterior=r_ant,
                            archivo_nuevo=r_nue,
                            hoja_anterior=self.hoja,
                            hoja_nueva=self.hoja,
                            columnas_clave=self.claves,
                            columnas_comparadas=self.comparar,
                            cambios=cambios,
                        )

                        total_c += len(cambios)
                        df_ant = df_nue

                    self.finalizado.emit(total_c, None)
                except Exception as ex:
                    self.finalizado.emit(0, ex)

        self.worker_multiple = WorkerMultiple(rutas, nombre_hoja, columnas_clave, columnas_comparar, ruta_bd)

        def _al_progresar(val, msg):
            dialogo.setValue(val)
            dialogo.setLabelText(msg)

        def _al_finalizar(total_cambios, error):
            dialogo.close()
            if error:
                QMessageBox.critical(
                    self,
                    "Error en comparación múltiple",
                    f"No se pudo completar el procesamiento:\n\n{error}",
                )
                return

            repositorio = RepositorioSQLite(ruta_bd)

            QMessageBox.information(
                self,
                "Comparación múltiple completada",
                (
                    "Se han procesado secuencialmente todos los archivos "
                    f"({len(rutas)} archivos, {len(rutas) - 1} comparaciones).\n\n"
                    f"Cambios totales guardados: {total_cambios}\n\n"
                    f"Base de datos:\n{ruta_bd}"
                ),
            )

            # Cargar inmediatamente en el Histórico
            self.ui.lineEditBaseHistorico.setText(ruta_bd)
            self.ruta_historico = ruta_bd
            self.repositorio_historico = repositorio

            cambios_historico = repositorio.obtener_historico_cambios()
            self._actualizar_identificadores_historico(cambios_historico)
            self.modelo_historico.actualizar(cambios_historico)
            self.limpiar_detalle_historico()

            self.ui.tabWidget.setCurrentWidget(self.ui.tab_historico)
            self._actualizar_estado_db(ruta_bd)

            self.ui.statusBar.showMessage(
                f"Procesados {len(rutas)} archivos secuencialmente."
            )

        dialogo.canceled.connect(lambda: self.worker_multiple.requestInterruption())
        self.worker_multiple.progreso.connect(_al_progresar)
        self.worker_multiple.finalizado.connect(_al_finalizar)
        self.worker_multiple.start()
